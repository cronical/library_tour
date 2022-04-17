import pandas as pd
from openpyxl import load_workbook
from numpy import nan
from geopy.geocoders import Nominatim
import geopy.extra.rate_limiter as rl
file_name='geocoded_libraries.xlsx'
ll='latitude'	,'longitude'
#using load_workbook and converting preserves the formulas
wb=load_workbook(file_name,data_only=False) 
ws=wb['Sheet1']
data=ws.values
columns=next(data)[0:]
df=pd.DataFrame(data,columns=columns)
geolocator = Nominatim(user_agent="library-tour")
geocoder = rl.RateLimiter(geolocator.geocode, min_delay_seconds=1)
lat=[]
long=[]
for ix,row in df.iterrows():
  print(f"{ix}. {row['Library']}")
  lat=nan
  long=nan
  location = geocoder(row['Full Address'])
  if location is not None:
    lat=location.latitude
    long=location.longitude
  df.loc[ix,'latitude']=lat
  df.loc[ix,'longitude']=long
print(df)
df.to_excel('geocoded_libraries.xlsx')